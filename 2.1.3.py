import csv
import re
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Border, Side
# D:/пользователи/OneDrive/Рабочий стол/vacancies.csv
from prettytable import PrettyTable, ALL
import matplotlib.pyplot as plt
import numpy as np
from jinja2 import Environment, FileSystemLoader
import pathlib
import pdfkit
#that is the string i am going to change!!!!!!
polina_name = 'polina'

requests = ["Введите название файла: ", "Введите название профессии: "]


class Vacancy:
    names = ['number', 'name', 'description', 'key_skills', 'experience_id', 'premium', 'employer_name', 'salary',
             'area_name', 'published_at']

    currency_to_rub = {
        "AZN": 35.68,
        "BYR": 23.91,
        "EUR": 59.90,
        "GEL": 21.74,
        "KGS": 0.76,
        "KZT": 0.13,
        "RUR": 1,
        "UAH": 1.64,
        "USD": 60.66,
        "UZS": 0.0055,
    }

    @staticmethod
    def format_string(x):
        str = re.sub(r'<.*?>', '', x)
        return re.sub(r'\s+', ' ', str).strip()

    @staticmethod
    def cut_the_str(str):
        if len(str) > 100:
            newstr = (str[:100] + '...')
        else:
            newstr = str
        return newstr

    @property
    def experience(self):
        return self.experience_items[self.experience_id]

    experience_items = {
        'Нет опыта': 0,
        'От 1 года до 3 лет': 1,
        'От 3 до 6 лет': 2,
        'Более 6 лет': 3
    }
    translate_currency = {
        'AZN': 'Манаты',
        'BYR': 'Белорусские рубли',
        'EUR': 'Евро',
        'GEL': 'Грузинский лари',
        'KGS': 'Киргизский сом',
        'KZT': 'Тенге',
        'RUR': 'Рубли',
        'UAH': 'Гривны',
        'USD': 'Доллары',
        'UZS': 'Узбекский сум'
    }
    translate_experience = {
        'noExperience': 'Нет опыта',
        'between1And3': 'От 1 года до 3 лет',
        'between3And6': 'От 3 до 6 лет',
        'moreThan6': 'Более 6 лет',
    }

    def make_array(self):
        result = []
        for key in self.names:
            result.append(getattr(self, key))
        return result

    def __init__(self, vacancy):
        self.name = vacancy['name']
        self.salary_from = int(vacancy['salary_from'].split('.')[0])
        self.salary_to = int(vacancy['salary_to'].split('.')[0])
        self.salary_currency = vacancy['salary_currency']
        self.average_salary = self.currency_to_rub[self.salary_currency] * (self.salary_from + self.salary_to) / 2
        self.area_name = vacancy['area_name']
        self.year = int(vacancy['published_at'].rsplit('T', 1)[0].split('-')[0])
        self.published = vacancy['published_at']
        arr = vacancy['published_at'].rsplit('T', 1)[0].replace('-', '.').split('.')
        arr.reverse()
        self.published_at = '.'.join(arr)
        if len(vacancy) > 6:
            self.description = self.cut_the_str(self.format_string(vacancy['description']))
            self.skills = vacancy['key_skills'].split('\n')
            self.key_skills = self.cut_the_str(vacancy['key_skills'])
            self.len_skills = len(self.skills)
            self.experience_id = self.translate_experience[vacancy['experience_id']]
            if vacancy['premium'].lower() == 'true':
                self.premium = 'Да'
            elif vacancy['premium'].lower() == 'false':
                self.premium = 'Нет'
            self.employer_name = vacancy['employer_name']
            if vacancy['salary_gross'].lower() == 'true':
                self.salary_gross = 'Без вычета налогов'
            else:
                self.salary_gross = 'С вычетом налогов'
            self.salary = '{0:,}'.format(int(float(self.salary_from))).replace(',', ' ') + ' - ' + \
                '{0:,}'.format(int(float(self.salary_to))).replace(',', ' ') + ' (' + self.translate_currency[self.salary_currency]  + ')' \
                + ' (' + self.salary_gross + ')'


class DataSet:
    def __init__(self, fileName, vacancyName, filterPar, sortPar, sortOrder, inDataNumbers):
        self.fileName = fileName
        self.vacancyName = vacancyName
        self.filterPar = filterPar
        self.sortPar = sortPar
        self.sortOrder = sortOrder
        self.inDataNumbers = inDataNumbers
        self.vacancies_objects = []

    translate_titles_dict = {'Название': 'name', 'Описание': 'description', 'Навыки': 'len_skills',
                             'Опыт работы': 'experience',
                             'Премиум-вакансия': 'premium', 'Компания': 'employer_name', 'Оклад': 'average_salary',
                             'Идентификатор валюты оклада': 'salary_currency', 'Название региона': 'area_name',
                             'Дата публикации вакансии': 'published'}

    def csv_reader2(self):
        r = open(self.fileName, encoding='utf-8-sig')
        file = csv.reader(r)

        titles = []
        text = [x for x in file]
        counter = 0
        spes = []
        for specialitiy in text:
            if counter == 0:
                titles = specialitiy
                l = len(specialitiy)
            else:
                if len([parameter for parameter in specialitiy if parameter]) == l:
                    spes.append(specialitiy)
            counter += 1

        newSpecialities = []
        for eachList in spes:
            newList = []
            for x in eachList:
                newList.append(x)

            newSpecialities.append(newList)

        for list1 in newSpecialities:
            newDict = dict(zip(titles, list1))  # cоздали словарь
            self.vacancies_objects.append(Vacancy(newDict))

        if len(self.vacancies_objects) == 0 and len(titles) == 0:
            print('Пустой файл')
            quit()
        elif len(self.vacancies_objects) == 0:
            print('Нет данных')
            quit()

    def csv_reader(self):
        r = open(self.fileName, encoding='utf-8-sig')
        file = csv.reader(r)
        titles = []
        for counter, specialitiy in enumerate(file):
            if counter == 0:
                titles = specialitiy
                l = len(specialitiy)
            else:
                if len([parameter for parameter in specialitiy if parameter]) == l:
                    yield dict(zip(titles, specialitiy))

    sort_params = {
        'Название': lambda vacancy, value: vacancy.name == value,
        'Навыки': lambda vacancy, value: all([skill in vacancy.skills for skill in value.split(', ')]),
        'Опыт работы': lambda vacancy, value: vacancy.experience_id == value,
        'Премиум-вакансия': lambda vacancy, value: vacancy.premium == value,
        'Компания': lambda vacancy, value: vacancy.employer_name == value,
        'Оклад': lambda vacancy, value: vacancy.salary_from <= float(value) <= vacancy.salary_to,
        'Идентификатор валюты оклада': lambda vacancy, value: vacancy.translate_currency[
                                                                  vacancy.salary_currency] == value,
        'Название региона': lambda vacancy, value: vacancy.area_name == value,
        'Дата публикации вакансии': lambda vacancy, value: vacancy.published_at == value
    }

    def get_the_rows_for_table(self):
        result = []
        for item in self.vacancies_objects:
            result.append(item.make_array())
        return result

    def filter_for_table_rows(self):
        if len(self.filterPar) < 1:
            return

        result = filter(lambda x: self.sort_params[self.filterPar[0]](x, self.filterPar[1]), self.vacancies_objects)
        self.vacancies_objects = list(result)

    def sort_for_table_rows(self):
        if self.sortPar == '' and self.sortOrder:
            self.vacancies_objects.reverse()

        elif self.sortPar != '':
            self.vacancies_objects.sort(key=lambda x: getattr(x, DataSet.translate_titles_dict[self.sortPar]),
                                        reverse=self.sortOrder)

    def make_numbers_to_sort(self):
        sort_order_len = len(self.inDataNumbers)
        array = []

        for number, item in enumerate(self.vacancies_objects):
            if (sort_order_len > 1 and self.inDataNumbers[0] <= number < self.inDataNumbers[1]) or (
                    sort_order_len == 1 and self.inDataNumbers[0] <= number) or sort_order_len == 0:
                item.number = number + 1
                array.append(item)
        self.vacancies_objects = array

    @staticmethod
    def check_key(dictionary, key, var):
        if key in dictionary:
            dictionary[key] += var
        else:
            dictionary[key] = var

    def make_statistic_data(self):
        salaryDict = {}
        salaryArea = {}
        vacancySalary = {}
        vacanciesCounter = 0

        for dictionar in self.csv_reader():
            vacancy = Vacancy(dictionar)
            self.check_key(salaryDict, vacancy.year, [vacancy.average_salary])
            if vacancy.name.find(self.vacancyName) != -1:
                self.check_key(vacancySalary, vacancy.year, [vacancy.average_salary])
            self.check_key(salaryArea, vacancy.area_name, [vacancy.average_salary])
            vacanciesCounter += 1

        arr = []
        for key, value in salaryDict.items():
            arr.append((key, len(value)))
        numberOfVacancies = dict(arr)

        arr1 = []
        for key, value in vacancySalary.items():
            arr1.append((key, len(value)))
        vacancyNameCounter = dict(arr1)

        if not vacancySalary:
            arr2 = []
            for key, value in salaryDict.items():
                arr2.append((key, [0]))
            vacancySalary = dict(arr2)

            arr3 = []
            for key, value in numberOfVacancies.items():
                arr3.append((key, 0))
            vacancyNameCounter = dict(arr3)

        statistics_of_salary = self.count_average_number(salaryDict)
        statistics_of_salary_by_vacancy = self.count_average_number(vacancySalary)
        statistics_of_salary_by_area = self.count_average_number(salaryArea)
        temp_dict_of_statistics = {}
        for publishedData, salariesArr in salaryArea.items():
            temp_dict_of_statistics[publishedData] = round(len(salariesArr) / vacanciesCounter, 4)
        temp_arr = []
        for key, value in temp_dict_of_statistics.items():
            temp_arr.append((key, value))
        filtered_arr = filter(lambda x: x[-1] >= 0.01, temp_arr)
        temp_dict_of_statistics = list(filtered_arr)
        temp_dict_of_statistics.sort(key=lambda x: x[-1], reverse=True)
        result_statistics = temp_dict_of_statistics
        temp_dict_of_statistics = dict(temp_dict_of_statistics)

        temp_arr_2 = []
        for key, value in statistics_of_salary_by_area.items():
            temp_arr_2.append((key, value))
        filtered_arr_2 = filter(lambda x: x[0] in list(temp_dict_of_statistics.keys()), temp_arr_2)
        statistics_of_salary_by_area = list(filtered_arr_2)
        statistics_of_salary_by_area.sort(key=lambda a: a[-1], reverse=True)
        statistics_of_salary_by_area = dict(statistics_of_salary_by_area[:10])
        result_statistics = dict(result_statistics[:10])

        return statistics_of_salary, numberOfVacancies, statistics_of_salary_by_vacancy, vacancyNameCounter, statistics_of_salary_by_area, result_statistics

    @staticmethod
    def count_average_number(dict_to_count):
        result = {}
        for key, value in dict_to_count.items():
            a = sum(value)
            b = len(value)
            result[key] = int(a / b)
        return result


class Report:
    def __init__(self, vacancyName, statistics, statistics_of_salary_by_vacancy, statistics_of_salary_by_area,
                 temp_dict_of_statistics, result_statistics, statistics_of_vacancy_by_year):
        self.wb = Workbook()
        self.vacancyName = vacancyName
        self.statistics = statistics
        self.statistics_of_salary_by_vacancy = statistics_of_salary_by_vacancy
        self.statistics_of_salary_by_area = statistics_of_salary_by_area
        self.temp_dict_of_statistics = temp_dict_of_statistics
        self.result_statistics = result_statistics
        self.statistics_of_vacancy_by_year = statistics_of_vacancy_by_year

    def generate_excel(self):
        list1 = self.wb.active
        list1.title = 'Статистика по годам'
        array_for_titles = ['Год', 'Средняя зарплата', 'Средняя зарплата - ' + self.vacancyName, 'Количество вакансий',
                            'Количество вакансий - ' + self.vacancyName]
        list1.append(array_for_titles)
        for date in self.statistics.keys():
            list1.append([date, self.statistics[date], self.statistics_of_salary_by_area[date],
                          self.statistics_of_salary_by_vacancy[date], self.temp_dict_of_statistics[date]])

        titles = [['Год ', 'Средняя зарплата ', ' Средняя зарплата - ' + self.vacancyName, ' Количество вакансий',
                   ' Количество вакансий - ' + self.vacancyName]]
        widthOfColums = []
        for number in titles:
            for i, element in enumerate(number):
                widthOfColums = self.set_colums_length(element, i, widthOfColums)

        for i, j in enumerate(widthOfColums, 1):
            converted_var = get_column_letter(i)
            list1.column_dimensions[converted_var].width = j + 2

        titles = []
        titles.append(['Город', 'Уровень зарплат', '', 'Город', 'Доля вакансий'])
        new_dictionary = zip(self.result_statistics.items(), self.statistics_of_vacancy_by_year.items())
        for (first_area, i), (second_area, j) in new_dictionary:
            titles.append([first_area, i, '', second_area, j])

        list2 = self.wb.create_sheet('Статистика по городам')
        for number in titles:
            list2.append(number)

        widthOfColums = []
        for number in titles:
            for i, element in enumerate(number):
                element = str(element)
                if len(widthOfColums) > i:
                    if len(element) > widthOfColums[i]:
                        widthOfColums[i] = len(element)
                else:
                    widthOfColums += [len(element)]

        for i, j in enumerate(widthOfColums, 1):
            list2.column_dimensions[get_column_letter(i)].width = j + 2

        bold = Font(bold=True)
        for column in 'ABCDE':
            list1[column + '1'].font = bold
            list2[column + '1'].font = bold

        for x, y in enumerate(self.result_statistics):
            string_for_x = str(x + 2)
            list2['E' + string_for_x].number_format = '0.00%'

        thin = Side(border_style='thin', color='00000000')

        for number in range(len(titles)):
            for column in 'ABDE':
                self.make_border(column, number, thin, list2)

        for number, y in enumerate(self.statistics):
            for column in 'ABCDE':
                self.make_border(column, number, thin, list1)

        self.wb.save('report.xlsx')

    def generate_pdf(self):
        env = Environment(loader=FileSystemLoader('.'))
        template = env.get_template("python.html")
        reault_array = []
        for date in self.statistics.keys():
            reault_array.append([date, self.statistics[date], self.statistics_of_salary_by_vacancy[date],
                                 self.statistics_of_salary_by_area[date], self.temp_dict_of_statistics[date]])
        for item in self.statistics_of_vacancy_by_year:
            rounded = round(self.statistics_of_vacancy_by_year[item] * 100, 2)
            self.statistics_of_vacancy_by_year[item] = rounded

        pdf_template = template.render(
            {'name': self.vacancyName, 'path': '{0}/{1}'.format(pathlib.Path(__file__).parent.resolve(), 'graph.png'),
             'reault_array': reault_array, 'result_statistics': self.result_statistics,
             'statistics_of_vacancy_by_year': self.statistics_of_vacancy_by_year})

        config = pdfkit.configuration(wkhtmltopdf=r'C:\Program Files\wkhtmltopdf\bin\wkhtmltopdf.exe')
        pdfkit.from_string(pdf_template, 'report.pdf', configuration=config, options={"enable-local-file-access": ""})

    def generate_image(self):
        fig, ((ax1, ax2), (ax3, ax4)) = plt.subplots(nrows=2, ncols=2)

        ax1.set_title('Уровень зарплат по годам', fontdict={'fontsize': 12})
        ax1.grid(axis='y')

        list_for_stats = list(self.statistics.keys())
        ax1.legend((ax1.bar(np.array(list_for_stats) - 0.4, self.statistics.values(), width=0.4)[0],
                    ax1.bar(np.array(list_for_stats), self.statistics_of_salary_by_area.values(), width=0.4)[0]),
                   ('средняя з/п', 'з/п ' + self.vacancyName.lower()), prop={'size': 8})
        ax1.set_xticks(np.array(list_for_stats) - 0.2, list_for_stats, rotation=90)
        ax1.xaxis.set_tick_params(labelsize=8)
        ax1.yaxis.set_tick_params(labelsize=8)

        list_for_salary_stats = list(self.statistics_of_salary_by_vacancy.keys())
        ax2.set_title('Количество вакансий по годам', fontdict={'fontsize': 12})
        ax2.legend((ax2.bar(np.array(list_for_salary_stats) - 0.4,
                            self.statistics_of_salary_by_vacancy.values(), width=0.4)[0],
                    ax2.bar(np.array(list_for_salary_stats), self.temp_dict_of_statistics.values(), width=0.4)[0]),
                   ('Количество вакансий', 'Количество вакансий\n' + self.vacancyName.lower()), prop={'size': 8})
        ax2.set_xticks(np.array(list_for_salary_stats) - 0.2, list_for_salary_stats, rotation=90)
        ax2.grid(axis='y')
        ax2.xaxis.set_tick_params(labelsize=8)
        ax2.yaxis.set_tick_params(labelsize=8)

        list_for_result_keys = list(self.result_statistics.keys())
        list_for_result_values = list(self.result_statistics.values())
        ax3.set_title('Уровень зарплат по городам', fontdict={'fontsize': 12})

        new_list = []
        for a in reversed(list_for_result_keys):
            new_list.append(str(a).replace(' ', '\n').replace('-', '-\n'))

        ax3.barh(new_list, list(reversed(list_for_result_values)), align='center')
        ax3.yaxis.set_tick_params(labelsize=6)
        ax3.xaxis.set_tick_params(labelsize=8)
        ax3.grid(axis='x')

        list_for_stats_date_values = list(self.statistics_of_vacancy_by_year.values())
        list_for_stats_date_keys = list(self.statistics_of_vacancy_by_year.keys())

        ax4.set_title('Доля вакансий по городам', fontdict={'fontsize': 12})

        counter = []
        for i in self.statistics_of_vacancy_by_year.values():
            counter.append(i)

        others = 1 - sum(counter)
        ax4.pie([others] + list_for_stats_date_values, labels=['Другие'] + list_for_stats_date_keys,
                textprops={'fontsize': 6})

        plt.tight_layout()
        plt.savefig('graph.png')

    def set_colums_length(self, element, i, widthOfColums):
        length = len(element)
        if len(widthOfColums) > i:
            if length > widthOfColums[i]:
                widthOfColums[i] = length
        else:
            widthOfColums += [length]
        return widthOfColums

    def make_border(self, column, number, thin, list2):
        list2[column + str(number + 1)].border = Border(left=thin, bottom=thin, right=thin, top=thin)


class Table:
    titles = ['№', 'Название', 'Описание', 'Навыки', 'Опыт работы', 'Премиум-вакансия', 'Компания', 'Оклад',
              'Название региона', 'Дата публикации вакансии']

    def __init__(self):
        self.incorrectOutputs = []
        self.fileName = input('Введите название файла: ')
        self.filterPar = self.filter_needed(input('Введите параметр фильтрации: '))
        self.sortPar = self.sortPar_needed(input('Введите параметр сортировки: '))
        self.sortOrder = self.sortOrder_needed(input('Обратный порядок сортировки (Да / Нет): '))
        self.inDataNumbers = self.in_data_numbers_needed(input('Введите диапазон вывода: '))
        self.tableColums = self.in_data_names_needed(input('Введите требуемые столбцы: '))

        if len(self.incorrectOutputs) > 0:
            print(self.incorrectOutputs[0])
            quit()

        data_set = DataSet(self.fileName,'', self.filterPar, self.sortPar, self.sortOrder, self.inDataNumbers)
        data_set.csv_reader2()
        data_set.filter_for_table_rows()
        data_set.sort_for_table_rows()
        data_set.make_numbers_to_sort()

        if len(data_set.get_the_rows_for_table()) == 0:
            print('Ничего не найдено')
        else:
            table = PrettyTable()
            table.hrules = ALL
            table.align = 'l'
            table.field_names = Table.titles
            table._max_width = {"№": 20, "Название": 20, "Описание": 20,
                                "Навыки": 20, "Опыт работы": 20, "Премиум-вакансия": 20,
                                "Компания": 20, "Оклад": 20, "Название региона": 20, "Дата публикации вакансии": 20}
            table.add_rows(data_set.get_the_rows_for_table())
            print(table.get_string(fields=self.tableColums))

    def filter_needed(self, filterPar):
        filter_is_needed = len(filterPar) > 1
        if not filter_is_needed:
            return []
        if filter_is_needed:
            if ':' not in filterPar:
                self.incorrectOutputs.append('Формат ввода некорректен')
                return []

            filterPar = filterPar.split(': ')
            if (not filterPar[0] in Table.titles) and filterPar[0] != 'Идентификатор валюты оклада':
                self.incorrectOutputs.append('Параметр поиска некорректен')
                return []

        return filterPar

    def sortPar_needed(self, sortPar):
        sortPar_is_needed = len(sortPar) > 1
        if sortPar_is_needed:
            if sortPar not in Table.titles:
                self.incorrectOutputs.append('Параметр сортировки некорректен')
        return sortPar

    def sortOrder_needed(self, sortOrder):
        sort_order_needed = len(str(sortOrder)) > 1
        if sort_order_needed:
            if sortOrder != 'Да' and sortOrder != 'Нет':
                self.incorrectOutputs.append('Порядок сортировки задан некорректно')
        if sortOrder == 'Да':
            return True
        else:
            return False

    def in_data_numbers_needed(self, inDataNumbers):
        arrOfinDataNumbers = []

        if inDataNumbers == '':
            return []
        else:
            for a in inDataNumbers.split(' '):
                arrOfinDataNumbers.append(int(a) - 1)
            return arrOfinDataNumbers

        return [] if sort_range == '' else [int(limit) - 1 for limit in sort_range.split()]

    def in_data_names_needed(self, tableColums):
        if len(tableColums) == 0:
            return Table.titles
        else:
            arrOfinDataNames = tableColums.split(', ')
            arrOfinDataNames.insert(0, '№')
            return arrOfinDataNames


class InputConnect:
    def __init__(self):
        output = False if input("(Статистика/Вакансии): ") == "Статистика" else True
        if output:
            Table()
        else:
            self.fileName = input(requests[0])
            self.vacancyName = input(requests[1])
            data_set = DataSet(self.fileName, self.vacancyName, '', '', '', '')
            statistics, statistics_of_salary_by_vacancy, statistics_of_salary_by_area, temp_dict_of_statistics, result_statistics, statistics_of_vacancy_by_year = data_set.make_statistic_data()

            print('Динамика уровня зарплат по годам: {0}'.format(statistics))
            print('Динамика количества вакансий по годам: {0}'.format(statistics_of_salary_by_vacancy))
            print('Динамика уровня зарплат по годам для выбранной профессии: {0}'.format(statistics_of_salary_by_area))
            print('Динамика количества вакансий по годам для выбранной профессии: {0}'.format(temp_dict_of_statistics))
            print('Уровень зарплат по городам (в порядке убывания): {0}'.format(result_statistics))
            print('Доля вакансий по городам (в порядке убывания): {0}'.format(statistics_of_vacancy_by_year))

            report = Report(self.vacancyName, statistics, statistics_of_salary_by_vacancy, statistics_of_salary_by_area,
                            temp_dict_of_statistics, result_statistics, statistics_of_vacancy_by_year)
            report.generate_excel()
            report.wb.save(filename='report.xlsx')
            report.generate_image()
            report.generate_pdf()


if __name__ == '__main__':
    InputConnect()
